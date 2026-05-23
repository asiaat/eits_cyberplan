"""IMR Service - E-ITS Implementation Plan Management."""
import uuid
import datetime
from typing import Dict, Any, List, Optional
from sqlalchemy.orm import Session
from fastapi import HTTPException, status

from app.models import (
    ImrItem, EvidenceLink, LocalUser, Tenant, 
    ProtectionNeedSummary, EitsCatalogMeasure
)
from app.models.eits_catalog_measure import EitsCatalogMeasure


class ImrService:

    @staticmethod
    def validate_imr_transition_to_implemented(
        db: Session,
        imr_item: ImrItem
    ) -> Dict[str, Any]:
        """
        Validate that an IMR item can transition to Implemented (R) status.
        E-ITS Business Rule: Measure cannot transition to Implemented (R) unless:
        1. implementation_details contains comprehensive explanation (min 15 chars)
        2. At least one active Evidence record is linked via evidence_links table
        """
        validation_errors = []
        
        # Check implementation_details
        if not imr_item.implementation_description or len(imr_item.implementation_description.strip()) < 15:
            validation_errors.append(
                "Tõrge: Meedet ei saa märkida rakendatuks (R) ilma piisava teostuskirjelduseta (min 15 tähemärki)."
            )
        
        # Check evidence links
        has_evidence = db.query(EvidenceLink).filter(
            EvidenceLink.target_id == imr_item.id,
            EvidenceLink.target_type == "imr_item",
            EvidenceLink.tenant_id == imr_item.tenant_id
        ).count() > 0
        
        if not has_evidence:
            validation_errors.append(
                "Tõrge: Rakendatud (R) meede nõuab vähemalt ühe digitaalse asitõendi seostamist."
            )
        
        return {
            "is_valid": len(validation_errors) == 0,
            "errors": validation_errors
        }

    @staticmethod
    def update_imr_item(
        db: Session,
        tenant_id: uuid.UUID,
        user_id: uuid.UUID,
        imr_item_id: uuid.UUID,
        update_data: Dict[str, Any]
    ) -> Dict[str, Any]:
        """
        Update an IMR item's implementation parameters.
        Enforces PEARO status transition rules.
        """
        # Get the IMR item
        item = db.query(ImrItem).filter(
            ImrItem.id == imr_item_id,
            ImrItem.tenant_id == tenant_id,
        ).first()
        
        if not item:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Seda rakenduskava meedet ei leitud."
            )
        
        # Check if status is being changed to 'R' (Implemented)
        new_status = update_data.get("status")
        if new_status == "R" or new_status == "Implemented":
            # Validate transition to Implemented status
            validation = ImrService.validate_imr_transition_to_implemented(db, item)
            if not validation["is_valid"]:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="; ".join(validation["errors"])
                )
        
        # Track status change
        if new_status and new_status != item.status:
            item.status_changed_at = datetime.datetime.utcnow()
        
        # Apply updates
        for key, value in update_data.items():
            if hasattr(item, key):
                setattr(item, key, value)
        
        item.updated_by = user_id
        item.updated_at = datetime.datetime.utcnow()
        
        db.commit()
        db.refresh(item)
        
        return {
            "message": "Rakenduskava meede edukalt uuendatud.",
            "imr_item_id": item.id,
            "status": item.status
        }

    @staticmethod
    def link_evidence_to_imr(
        db: Session,
        tenant_id: uuid.UUID,
        user_id: uuid.UUID,
        imr_item_id: uuid.UUID,
        evidence_id: uuid.UUID
    ) -> Dict[str, Any]:
        """
        Link evidence to an IMR item via polymorphic evidence_links table.
        """
        # Verify IMR item exists and belongs to tenant
        imr_item = db.query(ImrItem).filter(
            ImrItem.id == imr_item_id,
            ImrItem.tenant_id == tenant_id,
        ).first()
        
        if not imr_item:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="IMR item not found"
            )
        
        # Verify evidence exists and belongs to tenant
        from app.models.evidence import Evidence
        evidence = db.query(Evidence).filter(
            Evidence.id == evidence_id,
            Evidence.tenant_id == tenant_id,
        ).first()
        
        if not evidence:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Evidence not found"
            )
        
        # Check if link already exists
        existing_link = db.query(EvidenceLink).filter(
            EvidenceLink.tenant_id == tenant_id,
            EvidenceLink.evidence_id == evidence_id,
            EvidenceLink.target_id == imr_item_id,
            EvidenceLink.target_type == "imr_item"
        ).first()
        
        if existing_link:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Evidence is already linked to this IMR item"
            )
        
        # Create the evidence link
        evidence_link = EvidenceLink(
            tenant_id=tenant_id,
            evidence_id=evidence_id,
            target_id=imr_item_id,
            target_type="imr_item"
        )
        
        db.add(evidence_link)
        db.commit()
        
        return {
            "message": "Evidence successfully linked to IMR item",
            "evidence_link_id": evidence_link.id
        }

    @staticmethod
    def get_imr_item_with_validation_status(
        db: Session,
        tenant_id: uuid.UUID,
        imr_item_id: uuid.UUID
    ) -> Dict[str, Any]:
        """
        Get IMR item with validation status for Implemented (R) transition.
        """
        item = db.query(ImrItem).filter(
            ImrItem.id == imr_item_id,
            ImrItem.tenant_id == tenant_id,
        ).first()
        
        if not item:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="IMR item not found"
            )
        
        # Get validation status
        validation = ImrService.validate_imr_transition_to_implemented(db, item)
        
        # Get linked evidence count
        evidence_count = db.query(EvidenceLink).filter(
            EvidenceLink.target_id == imr_item_id,
            EvidenceLink.target_type == "imr_item",
            EvidenceLink.tenant_id == tenant_id
        ).count()
        
        return {
            "imr_item": item,
            "can_transition_to_implemented": validation["is_valid"],
            "validation_errors": validation["errors"],
            "linked_evidence_count": evidence_count,
            "has_sufficient_implementation_details": bool(
                item.implementation_description and 
                len(item.implementation_description.strip()) >= 15
            )
        }

    @staticmethod
    def bulk_update_imr_status(
        db: Session,
        tenant_id: uuid.UUID,
        user_id: uuid.UUID,
        imr_item_ids: List[uuid.UUID],
        new_status: str
    ) -> Dict[str, Any]:
        """
        Bulk update PEARO status for multiple IMR items with validation.
        """
        if new_status == "R":
            # For bulk R status, validate each item individually
            validation_errors = []
            valid_items = []
            
            for imr_item_id in imr_item_ids:
                item = db.query(ImrItem).filter(
                    ImrItem.id == imr_item_id,
                    ImrItem.tenant_id == tenant_id,
                ).first()
                
                if not item:
                    validation_errors.append(f"IMR item {imr_item_id} not found")
                    continue
                
                validation = ImrService.validate_imr_transition_to_implemented(db, item)
                if not validation["is_valid"]:
                    validation_errors.extend([
                        f"IMR item {imr_item_id}: {error}" 
                        for error in validation["errors"]
                    ])
                else:
                    valid_items.append(imr_item_id)
            
            if validation_errors:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="; ".join(validation_errors)
                )
            
            # Update valid items
            items = db.query(ImrItem).filter(
                ImrItem.id.in_(valid_items),
                ImrItem.tenant_id == tenant_id,
            ).all()
        else:
            # For non-R statuses, update all items
            items = db.query(ImrItem).filter(
                ImrItem.id.in_(imr_item_ids),
                ImrItem.tenant_id == tenant_id,
            ).all()
        
        # Update items
        for item in items:
            item.pearo_status = new_status
            item.updated_by = user_id
            item.updated_at = datetime.datetime.utcnow()
        
        db.commit()
        
        return {
            "message": f"Updated {len(items)} IMR items to status '{new_status}'",
            "updated_count": len(items)
        }

    @staticmethod
    def get_imr_summary_statistics(
        db: Session,
        tenant_id: uuid.UUID
    ) -> Dict[str, Any]:
        """
        Get IMR summary statistics for dashboard.
        """
        from sqlalchemy import func, case
        
        # Get counts by PEARO status
        status_counts = db.query(
            ImrItem.pearo_status,
            func.count(ImrItem.id).label('count')
        ).filter(
            ImrItem.tenant_id == tenant_id
        ).group_by(
            ImrItem.pearo_status
        ).all()
        
        # Convert to dict
        status_dict = {status: count for status, count in status_counts}
        
        # Get counts by priority
        priority_counts = db.query(
            ImrItem.priority,
            func.count(ImrItem.id).label('count')
        ).filter(
            ImrItem.tenant_id == tenant_id
        ).group_by(
            ImrItem.priority
        ).all()
        
        priority_dict = {p: count for p, count in priority_counts}
        
        # Get overdue items (status E or O and due_date < today)
        from datetime import date
        overdue_count = db.query(func.count(ImrItem.id)).filter(
            ImrItem.tenant_id == tenant_id,
            ImrItem.pearo_status.in_(["E", "O"]),
            ImrItem.due_date < date.today()
        ).scalar()
        
        # Get items ready for completion (have evidence + sufficient details but not R)
        ready_for_completion = db.query(func.count(ImrItem.id)).filter(
            ImrItem.tenant_id == tenant_id,
            ImrItem.pearo_status.in_(["P", "E"]),
            ImrItem.implementation_description.is_not(None),
            func.length(ImrItem.implementation_description) >= 15,
            # Has at least one evidence link
            db.query(EvidenceLink.id).filter(
                EvidenceLink.target_id == ImrItem.id,
                EvidenceLink.target_type == "imr_item",
                EvidenceLink.tenant_id == tenant_id
            ).exists()
        ).scalar()
        
        return {
            "pearo_status_counts": status_dict,
            "priority_counts": priority_dict,
            "overdue_count": overdue_count,
            "ready_for_completion_count": ready_for_completion,
            "total_items": sum(status_dict.values())
        }

    @staticmethod
    def export_imr_to_excel(
        db: Session,
        tenant_id: uuid.UUID,
        pearo_status: Optional[str] = None,
        priority: Optional[str] = None,
        overdue_only: bool = False,
    ) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        query = db.query(ImrItem).filter(ImrItem.tenant_id == tenant_id)
        if pearo_status:
            query = query.filter(ImrItem.pearo_status == pearo_status)
        if priority:
            query = query.filter(ImrItem.priority == priority)
        if overdue_only:
            from datetime import date
            query = query.filter(
                ImrItem.due_date < date.today(),
                ImrItem.pearo_status.in_(["E", "O"]),
            )
        items = query.order_by(ImrItem.created_at.desc()).all()

        peero_labels = {"P": "Pole asjakohane", "E": "Ei ole rakendatud", "A": "Aktsepteeritud risk", "R": "Rakendatud", "O": "Osaliselt"}
        priority_labels = {"P1": "Kõrge", "P2": "Keskmine", "P3": "Madal"}

        wb = Workbook()
        ws = wb.active
        ws.title = "IMR"

        headers = ["Kood", "Meede", "Staatus", "Prioriteet", "Tähtaeg", "Vastutaja", "Profiil", "Tegevused (TODO)", "Maksumus (EUR)", "Kirjeldus", "Verifitseerimine", "Järgmine ülevaade"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, item in enumerate(items, 2):
            measure = db.query(EitsCatalogMeasure).filter(EitsCatalogMeasure.id == item.measure_id).first()
            user_name = ""
            if item.responsible_user_id:
                user = db.query(LocalUser).filter(LocalUser.id == item.responsible_user_id).first()
                if user:
                    user_name = user.full_name

            profile = item.requirement_profile
            if not profile and measure:
                profile = "PÕHIMEEDE" if measure.measure_level == "BASE" else "PIIRATULT"

            row_data = [
                measure.code if measure else "",
                measure.name if measure else "",
                peero_labels.get(item.pearo_status, item.pearo_status),
                priority_labels.get(item.priority, item.priority),
                str(item.due_date) if item.due_date else "",
                user_name,
                profile or "",
                item.todo_description or "",
                float(item.cost_eur) if item.cost_eur else "",
                item.implementation_description or "",
                item.verification_method or "",
                str(item.next_review_date) if item.next_review_date else "",
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 24
        ws.column_dimensions["G"].width = 14
        ws.column_dimensions["H"].width = 40
        ws.column_dimensions["I"].width = 16
        ws.column_dimensions["J"].width = 50
        ws.column_dimensions["K"].width = 30
        ws.column_dimensions["L"].width = 16

        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()